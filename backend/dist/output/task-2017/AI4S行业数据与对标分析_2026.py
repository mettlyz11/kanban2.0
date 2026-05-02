#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI4S行业数据与对标分析数据表
生成Excel格式的数据分析报告
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel_report():
    """创建完整的Excel分析报告"""
    
    wb = Workbook()
    
    # 设置样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== Sheet 1: 市场规模预测 ====================
    ws1 = wb.active
    ws1.title = "市场规模预测"
    
    market_data = {
        '年份': ['2023', '2024', '2025E', '2026E', '2027E', '2028E', '2029E', '2030E', '2031E', '2032E'],
        '全球市场规模(亿美元)': [21.5, 30.2, 45.38, 62.5, 85.0, 112.0, 145.0, 182.0, 220.0, 262.3],
        '中国市场规模(亿元)': [45, 68, 128, 195, 290, 410, 550, 680, 765, 850],
        '全球增长率(%)': [np.nan, 40.5, 50.3, 37.7, 36.0, 31.8, 29.5, 25.5, 20.9, 19.2],
        '中国增长率(%)': [np.nan, 51.1, 88.2, 52.3, 48.7, 41.4, 34.1, 23.6, 12.5, 11.1]
    }
    
    df1 = pd.DataFrame(market_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df1, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    # 调整列宽
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 16
    
    # ==================== Sheet 2: 标杆企业财务对比 ====================
    ws2 = wb.create_sheet("标杆企业财务对比")
    
    finance_data = {
        '企业': ['晶泰控股', '英矽智能', 'Schrödinger', '深势科技(估)', 'Recursion'],
        '2025年收入': ['8.03亿元', '5624万美元', '3.82亿美元', '~3亿元', '2.1亿美元'],
        '收入同比': ['+201.2%', '+45%', '+35%', '+80%', '+28%'],
        '净利润': ['1.35亿元', '-1.2亿美元', '-1.12亿美元', '-1.5亿元', '-1.8亿美元'],
        '现金储备': ['70.69亿元', '3.93亿美元', '8.75亿美元', '~15亿元', '5.2亿美元'],
        '员工人数': ['~1500', '~800', '~1200', '~500', '~450'],
        'ARPU': ['~53万元', '~70万元', '~32万美元', '~60万元', '~47万美元'],
        '状态': ['已盈利', '未盈利', '未盈利', '未盈利', '未盈利']
    }
    
    df2 = pd.DataFrame(finance_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df2, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            # 高亮已盈利企业
            if r_idx > 1 and c_idx == 8 and value == '已盈利':
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws2.column_dimensions[col].width = 18
    
    # ==================== Sheet 3: 商业模式矩阵 ====================
    ws3 = wb.create_sheet("商业模式矩阵")
    
    biz_model_data = {
        '模式类型': ['AI+CRO服务', 'AI+Biotech管线', 'Science-as-a-Service', '垂直SaaS', '软件授权'],
        '代表企业': ['晶泰控股', '英矽智能', '深势科技', 'Rodin', 'Schrödinger'],
        '收入模式': ['服务费+里程碑+分成', '软件+里程碑+分成', '订阅+解决方案', 'SaaS订阅', '永久授权+维护'],
        '客单价范围': ['50万-10亿元', '10万-27.5亿美元', '10万-1000万', '$49-$199/月', '$5k-$50万/年'],
        '盈利周期': ['7-8年', '10年+', '8-10年', '2-3年', '5-7年'],
        '技术壁垒': ['★★★★★', '★★★★☆', '★★★★★', '★★★☆☆', '★★★★☆'],
        '可复制性': ['★★★☆☆', '★★☆☆☆', '★★★★☆', '★★★★★', '★★★☆☆'],
        '天花板': ['★★★★★', '★★★★★', '★★★★★', '★★★☆☆', '★★★☆☆'],
        '典型ARR增速': ['100-200%', '30-50%', '50-80%', '300%+', '15-25%']
    }
    
    df3 = pd.DataFrame(biz_model_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df3, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws3.column_dimensions[col].width = 18
    
    # ==================== Sheet 4: 商业化里程碑路径 ====================
    ws4 = wb.create_sheet("商业化里程碑路径")
    
    milestone_data = {
        '里程碑': ['技术验证', 'PMF达成', '百万ARR', '千万ARR', '盈利拐点'],
        '典型特征': [
            '核心算法SOTA，论文发表',
            '付费客户验证，NPS>40，留存>80%',
            '年经常性收入>100万美元，销售体系初步建立',
            '年经常性收入>1000万美元，商业模式验证',
            '经调整EBITDA转正，可持续经营'
        ],
        '平均耗时(年)': [1.5, 1.5, 1, 2, 3],
        'ARR区间': ['<100万', '100-500万', '100-500万', '1000-5000万', '>5000万'],
        '晶泰耗时': ['2年', '2年', '1年', '2年', '3年'],
        '英矽耗时': ['2年', '3年', '1年', '3年', '进行中'],
        '深势耗时': ['2年', '2年', '1年', '2年', '进行中'],
        '核心任务': [
            '技术可行性验证',
            '产品市场契合',
            '规模化销售验证',
            '商业模式验证',
            '可持续盈利能力'
        ]
    }
    
    df4 = pd.DataFrame(milestone_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df4, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    ws4.column_dimensions['A'].width = 15
    ws4.column_dimensions['B'].width = 45
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 15
    ws4.column_dimensions['E'].width = 12
    ws4.column_dimensions['F'].width = 12
    ws4.column_dimensions['G'].width = 12
    ws4.column_dimensions['H'].width = 25
    
    # ==================== Sheet 5: 核心成功要素对标 ====================
    ws5 = wb.create_sheet("核心成功要素对标")
    
    success_factor_data = {
        '成功要素': ['技术闭环能力', '跨学科团队', '高质量数据', '客户粘性', '里程碑变现'],
        '权重': ['30%', '20%', '20%', '15%', '15%'],
        '关键指标': [
            '数据→模型→客户→数据飞轮完整性',
            '博士占比、领域专家数量',
            '专有数据量、数据质量评分',
            '续费率、NPS、客户LTV',
            '里程碑金额/ARR比率'
        ],
        '晶泰评分': [9, 8, 9, 8, 10],
        '英矽评分': [8, 8, 8, 7, 9],
        '深势评分': [8, 9, 8, 8, 7],
        '和光目标值': [8, 8, 8, 8, 8],
        '差距分析': ['需加强实验层建设', '招聘/培养跨界人才', '建设数据平台', '提升客户成功体系', '设计里程碑机制']
    }
    
    df5 = pd.DataFrame(success_factor_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df5, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            # 评分列颜色标识
            if r_idx > 1 and c_idx in [4, 5, 6, 7]:
                if isinstance(value, (int, float)):
                    if value >= 9:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif value >= 7:
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws5.column_dimensions[col].width = 20
    
    # ==================== Sheet 6: 和光智成3年行动计划 ====================
    ws6 = wb.create_sheet("和光3年行动计划")
    
    action_data = {
        '行动领域': ['技术闭环', '技术闭环', '技术闭环', '变现路径', '变现路径', '变现路径', '生态建设', '生态建设', '生态建设'],
        '具体行动': [
            '实验自动化基础建设',
            '数据平台与AI模型迭代',
            '全闭环系统上线',
            '标准化SaaS平台上线',
            '大客户解决方案团队建设',
            '里程碑合作机制设计',
            '开源社区建设',
            '学术联盟建立',
            '产业投资布局'
        ],
        '时间周期': [
            '0-12个月',
            '12-24个月',
            '24-36个月',
            '0-12个月',
            '6-18个月',
            '12-24个月',
            '0-12个月',
            '12-24个月',
            '24-36个月'
        ],
        '预算(万元)': [2000, 3000, 5000, 500, 1500, 1000, 300, 800, 5000],
        '关键里程碑': [
            '3个核心实验流程自动化',
            '统一数据平台建成',
            '7×24无人实验室运行',
            '首批50家付费客户',
            '首个百万级项目交付',
            '首个千万级里程碑签约',
            '开源社区1000+开发者',
            '5家联合实验室建立',
            '3-5家生态企业投资'
        ],
        '负责人': [
            'CTO+研发VP',
            'CTO+数据总监',
            'CTO+实验室总监',
            '产品总监',
            '销售VP',
            'CEO+BD总监',
            '开源负责人',
            '战略合作总监',
            'CEO+CFO'
        ],
        '成功衡量指标': [
            '实验效率提升30%',
            '模型性能年提升20%',
            '数据自循环率>80%',
            '付费转化率>8%',
            '大客户签约率>20%',
            '里程碑收入占比>30%',
            '月活开发者>500',
            '联合论文>5篇/年',
            '投资企业>3家'
        ]
    }
    
    df6 = pd.DataFrame(action_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df6, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws6.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    ws6.column_dimensions['A'].width = 15
    ws6.column_dimensions['B'].width = 30
    ws6.column_dimensions['C'].width = 15
    ws6.column_dimensions['D'].width = 15
    ws6.column_dimensions['E'].width = 30
    ws6.column_dimensions['F'].width = 20
    ws6.column_dimensions['G'].width = 25
    
    # ==================== Sheet 7: 风险评估矩阵 ====================
    ws7 = wb.create_sheet("风险评估矩阵")
    
    risk_data = {
        '风险类别': ['技术风险', '技术风险', '技术风险', '商业化风险', '商业化风险', '商业化风险', '人才风险', '人才风险', '财务风险'],
        '具体风险': [
            '模型泛化能力不足',
            'AI可解释性不足',
            '技术迭代被超越',
            '客户付费意愿不足',
            '销售周期过长',
            '竞争加剧价格战',
            '跨界人才稀缺',
            '核心团队流失',
            '现金流断裂风险'
        ],
        '发生概率': ['中', '高', '中', '高', '高', '中', '高', '中', '低'],
        '影响程度': ['高', '高', '高', '中', '中', '中', '高', '高', '极高'],
        '风险等级': ['橙色', '红色', '橙色', '黄色', '黄色', '黄色', '红色', '橙色', '红色'],
        '应对策略': [
            '增加真实实验数据比例',
            '开发可解释AI技术',
            '预留20%研发投入前沿',
            '打造ROI可量化产品',
            '分层客户策略长短结合',
            '聚焦垂直领域建立壁垒',
            '培养+引进+生态三管齐下',
            '股权激励+文化建设',
            '多层次变现尽早现金平衡'
        ],
        '责任人': ['CTO', '首席科学家', 'CTO', '产品VP', '销售VP', 'CEO', 'HRVP', 'CEO', 'CFO'],
        '监控频率': ['季度', '季度', '月度', '月度', '月度', '季度', '月度', '季度', '月度']
    }
    
    df7 = pd.DataFrame(risk_data)
    
    for r_idx, row in enumerate(dataframe_to_rows(df7, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws7.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            # 风险等级颜色
            if r_idx > 1 and c_idx == 5:
                if value == '红色':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif value == '橙色':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws7.column_dimensions[col].width = 20
    
    # 保存文件
    output_path = '/Users/mettlyz/.openclaw/workspace/output/task-2017/AI4S行业数据与对标分析_2026.xlsx'
    wb.save(output_path)
    print(f"✅ Excel报告已生成: {output_path}")
    
    # 同时导出CSV格式便于数据处理
    csv_path = '/Users/mettlyz/.openclaw/workspace/output/task-2017/AI4S行业数据汇总_2026.csv'
    all_data = pd.concat([
        df1.assign(Sheet='市场规模预测'),
        df2.assign(Sheet='标杆企业财务对比'),
        df3.assign(Sheet='商业模式矩阵'),
        df4.assign(Sheet='商业化里程碑路径'),
        df5.assign(Sheet='核心成功要素对标'),
        df6.assign(Sheet='和光3年行动计划'),
        df7.assign(Sheet='风险评估矩阵')
    ], ignore_index=True)
    all_data.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f"✅ CSV数据汇总已生成: {csv_path}")
    
    return output_path, csv_path

def generate_summary_stats():
    """生成关键统计数据摘要"""
    
    summary = {
        '市场规模': {
            '2025全球AI4S市场(亿美元)': 45.38,
            '2032预测全球规模(亿美元)': 262.3,
            '年复合增长率CAGR(%)': 28.9,
            '中国2025市场规模(亿元)': 128,
            '中国CAGR(%)': 35.2
        },
        '标杆企业关键指标': {
            '晶泰2025营收(亿元)': 8.03,
            '晶泰净利润(亿元)': 1.35,
            '晶泰营收同比(%)': 201.2,
            '晶泰客户增长(%)': 62,
            '英矽2025营收(万美元)': 5624,
            '英矽累计合作总额(亿美元)': 46,
            '深势C轮融资(亿元)': 8,
            '深势服务科学家数量(万)': 300
        },
        '商业化里程碑': {
            'AI4S企业平均实现千万ARR耗时(年)': 7,
            '晶泰从成立到盈利耗时(年)': 10,
            'Rodin实现百万ARR耗时(天)': 45,
            '行业平均研发费用率(%)': 55,
            '头部企业客户续费率(%)': 75
        },
        '和光智成目标': {
            '第1年目标ARR(万元)': 1000,
            '第2年目标ARR(万元)': 3000,
            '第3年目标ARR(万元)': 8000,
            '预计突破千万美元ARR时间(年)': 3,
            '预计实现盈利时间(年)': 5
        }
    }
    
    summary_path = '/Users/mettlyz/.openclaw/workspace/output/task-2017/关键指标摘要_2026.json'
    import json
    with open(summary_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"✅ 关键指标摘要已生成: {summary_path}")
    
    return summary

if __name__ == '__main__':
    print("正在生成AI4S行业数据分析报告...\n")
    excel_path, csv_path = create_excel_report()
    summary = generate_summary_stats()
    
    print("\n" + "="*60)
    print("📊 数据报告生成完成！")
    print("="*60)
    print(f"📈 市场规模: 2025年全球45.38亿美元，CAGR 28.9%")
    print(f"🏆 标杆企业: 晶泰控股2025年营收8.03亿元，首次盈利")
    print(f"🎯 和光目标: 3年突破千万美元ARR，5年实现盈利")
    print("="*60)
