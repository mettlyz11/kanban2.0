#!/usr/bin/env python3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 创建工作簿
wb = Workbook()

# 定义样式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_worksheet(ws, df):
    """样式化工作表"""
    # 设置列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width
    
    # 设置表头样式
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 设置数据单元格样式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
    
    # 冻结首行
    ws.freeze_panes = "A2"

# ========== Sheet 1: 案例公司基本信息表 ==========
ws1 = wb.active
ws1.title = "案例公司基本信息"

basic_info = [
    ["公司名称", "成立时间", "总部", "上市时间/交易所", "累计融资", "估值/市值", "2024-2025营收", "员工规模", "数据来源"],
    ["Benchling", "2012年", "旧金山", "未上市（私有）", ">$8亿", "$61亿（2021年）", "$2.1亿 ARR（2024）", "~1200人", "Crunchbase, 投资者报告"],
    ["Schrödinger", "1990年", "纽约", "2020年NASDAQ (SDGR)", "IPO前~$3亿", "~$40亿（2026.04）", "$2.559亿（2025）", "~1500人", "公司财报, Yahoo Finance"],
    ["Insilico Medicine", "2014年", "香港/纽约", "2025年港交所 (3696.HK)", ">$10亿", "~$13亿（IPO）", "$5620万（2025）", "~500人", "公司年报, HKEX披露"],
    ["Recursion Pharmaceuticals", "2013年", "盐湖城", "2021年NASDAQ (RXRX)", "IPO后累计融资", "~$30亿（2026.04）", "$7468万（2025）", "~800人", "公司财报, 公开信息"],
    ["Citrine Informatics", "2013年", "红木城, 加州", "未上市", "$8130万", "$1.4亿（2022）", "$1210万（2024）", "68人", "GetLatka, Crunchbase"],
    ["Materials Zone", "2018年", "以色列特拉维夫", "未上市", "A轮（未公开）", "未公开", "接近千万ARR", "~50人", "公司官网, 行业报告"],
]

df1 = pd.DataFrame(basic_info[1:], columns=basic_info[0])
for r in dataframe_to_rows(df1, index=False, header=True):
    ws1.append(r)
style_worksheet(ws1, df1)

# ========== Sheet 2: 商业模式对比表 ==========
ws2 = wb.create_sheet("商业模式对比")

business_model = [
    ["公司名称", "核心商业模式", "SaaS订阅占比", "服务收入占比", "里程碑收入占比", "定价方式", "毛利率", "NRR/留存率", "客户数量"],
    ["Benchling", "纯SaaS订阅（RaaS）", "~95%", "~5%", "0%", "按席位+模块", "80%+", ">120%", "1200+"],
    ["Schrödinger", "软件+药物管线混合", "~60%", "~10%", "~30%", "订阅+永久授权+里程碑", "74%（软件）", "100%（大客户）", "Top 20药企全覆盖"],
    ["Insilico Medicine", "AI平台+管线双引擎", "~9%", "~44%", "~47%", "订阅+项目+授权", "81.5%", "N/A", "13家Top 20药企"],
    ["Recursion Pharmaceuticals", "合作里程碑为主", "~5%", "~10%", "~85%", "合作预付款+里程碑", "~50%", "N/A", "~10家大型药企"],
    ["Citrine Informatics", "纯SaaS订阅", "~90%", "~10%", "0%", "按模块+用户+数据量", "70%+", "N/A", "~50家企业"],
    ["Materials Zone", "SaaS平台模式", "~85%", "~15%", "0%", "按用户+数据规模", "N/A", "N/A", "~30家全球企业"],
]

df2 = pd.DataFrame(business_model[1:], columns=business_model[0])
for r in dataframe_to_rows(df2, index=False, header=True):
    ws2.append(r)
style_worksheet(ws2, df2)

# ========== Sheet 3: ARR里程碑时间线 ==========
ws3 = wb.create_sheet("ARR里程碑时间线")

milestone_data = [
    ["公司名称", "0→$1M ARR耗时", "$1M→$5M ARR耗时", "$5M→$10M ARR耗时", "达到$10M时间", "达到$100M时间", "总耗时（到$10M）", "关键加速因素"],
    ["Benchling", "~36个月", "~24个月", "~12个月", "~2018年", "~2021年", "~6年", "生命科学PMF精准+平台化"],
    ["Schrödinger", "~15年（传统软件）", "~5年", "~3年", "~2013年", "~2020年", "~23年", "30年技术积累+云端转型"],
    ["Insilico Medicine", "~5年", "~2年", "进行中", "预计2026年", "N/A", "预计~12年", "端到端平台+临床验证"],
    ["Recursion Pharmaceuticals", "~5年", "~3年", "~2年", "~2022年", "进行中", "~9年", "规模化湿实验+大额药企合作"],
    ["Citrine Informatics", "~60个月", "~24个月", "~24个月", "2024年", "N/A", "~11年", "材料垂直深耕+企业级产品"],
    ["Materials Zone", "预计36-48个月", "进行中", "未达到", "预计2026-2027", "N/A", "预计~8年", "产品易用性+端到端覆盖"],
    ["和光智成（目标）", "18个月", "12个月", "6个月", "2029年", "2031年", "~3年", "中国市场红利+聚焦材料AI"],
]

df3 = pd.DataFrame(milestone_data[1:], columns=milestone_data[0])
for r in dataframe_to_rows(df3, index=False, header=True):
    ws3.append(r)
style_worksheet(ws3, df3)

# ========== Sheet 4: 关键财务指标对比 ==========
ws4 = wb.create_sheet("关键财务指标对比")

financial_data = [
    ["公司名称", "最新年营收", "YoY增长率", "毛利率", "LTV/CAC估算", "CAC回收周期", "年客户流失率", "客单价范围", "销售周期", "现金储备"],
    ["Benchling", "$2.1亿（2024 ARR）", "27%", "80%+", "~5x", "~6个月", "<5%", "$10万-$100万+", "3-12个月", "N/A"],
    ["Schrödinger", "$2.559亿（2025）", "23.3%", "74%（软件）", "~5x", "~9个月", "0%（大客户）", "$50万-$500万+", "6-18个月", "~$4亿"],
    ["Insilico Medicine", "$5620万（2025）", "-34.5%", "81.5%", "~3x", "~12个月", "N/A", "$百万级合作", "9-18个月", "$3.93亿"],
    ["Recursion Pharmaceuticals", "$7468万（2025）", "27%", "~50%", "~4x", "~18个月", "N/A", "$千万级里程碑", "12-24个月", "N/A"],
    ["Citrine Informatics", "$1210万（2024）", "~30%", "70%+", "~4x", "~6个月", "<10%", "$10万-$50万", "6-12个月", "N/A"],
    ["Materials Zone", "接近千万", "~100%", "N/A", "N/A", "N/A", "N/A", "$5万-$30万", "3-9个月", "N/A"],
    ["SaaS行业优秀标准", "-", ">30%", ">75%", ">3x", "<12个月", "<5%", "-", "-", "-"],
    ["AI4S行业典型值", "-", "20-50%", "60-80%", "2.5x-5x", "6-18个月", "3-10%", "-", "-", "-"],
]

df4 = pd.DataFrame(financial_data[1:], columns=financial_data[0])
for r in dataframe_to_rows(df4, index=False, header=True):
    ws4.append(r)
style_worksheet(ws4, df4)

# ========== Sheet 5: 客户画像对比 ==========
ws5 = wb.create_sheet("客户画像对比")

customer_data = [
    ["公司名称", "主要客户行业", "代表客户", "客户集中度", "Top客户占比", "客户获取方式", "销售团队规模"],
    ["Benchling", "生物制药60%、生物技术30%、学术10%", "Pfizer, Novartis, Gilead", "相对分散", "Top 20药企70%+覆盖", "直销+行业会议", "~200人"],
    ["Schrödinger", "制药、材料、学术", "Top 20药企全覆盖", "相对集中", "Top客户收入占比高", "直销+合作伙伴", "~150人"],
    ["Insilico Medicine", "跨国药企、中国Biotech", "Eli Lilly, Servier, 齐鲁制药", "高度集中", "Top 5占94.4%（2024）", "BD团队+高层对接", "~50人"],
    ["Recursion Pharmaceuticals", "大型药企", "Roche, Genentech, Sanofi", "高度集中", "Top 2占收入主要部分", "战略合作BD", "~80人"],
    ["Citrine Informatics", "特种化工、先进材料、能源", "财富500强化工企业", "中度集中", "N/A", "直销+渠道伙伴", "~15人"],
    ["Materials Zone", "化工、建筑材料、特种材料", "Imerys, SGL Carbon, Cemex", "中度分散", "N/A", "直销+行业网络", "~10人"],
]

df5 = pd.DataFrame(customer_data[1:], columns=customer_data[0])
for r in dataframe_to_rows(df5, index=False, header=True):
    ws5.append(r)
style_worksheet(ws5, df5)

# ========== Sheet 6: 和光智成里程碑规划 ==========
ws6 = wb.create_sheet("和光智成里程碑规划")

heguang_data = [
    ["时间阶段", "营收里程碑", "付费客户数目标", "关键业务指标", "组织/产品里程碑", "核心任务"],
    ["Month 0-6", "$0 → $100K ARR", "3个POC客户", "POC成功率100%", "MVP产品上线, 核心团队到位", "技术验证, 寻找首批客户"],
    ["Month 6-12", "$100K → $500K ARR", "5-8个付费客户", "客单价$6万-$10万, 留存率100%", "完成PMF验证, 第1个标杆案例", "产品迭代, 建立POC流程"],
    ["Month 12-18", "$500K → $1M ARR", "10-15个付费客户", "月新增≥2个, NRR≥110%", "销售团队3-5人, 建立客户成功", "销售标准化, 验证增长模式"],
    ["Month 18-24", "$1M → $3M ARR", "25-35个付费客户", "月增长≥12%, LTV/CAC≥3x", "销售方法论沉淀, 完成A轮", "团队扩张, 品牌初步建立"],
    ["Month 24-30", "$3M → $6M ARR", "40-60个付费客户", "SaaS占比≥70%, 毛利率≥70%", "平台化架构, 跨行业验证", "规模化销售, 产品平台化"],
    ["Month 30-36", "$6M → $10M+ ARR", "70-100个付费客户", "年增长≥70%, NRR≥120%", "千万ARR里程碑, B轮准备", "生态建设, 寻求行业领导地位"],
]

df6 = pd.DataFrame(heguang_data[1:], columns=heguang_data[0])
for r in dataframe_to_rows(df6, index=False, header=True):
    ws6.append(r)
style_worksheet(ws6, df6)

# ========== Sheet 7: 风险应对矩阵 ==========
ws7 = wb.create_sheet("风险应对矩阵")

risk_data = [
    ["风险类型", "风险描述", "典型案例", "发生概率", "影响程度", "风险等级", "应对策略", "监控指标"],
    ["客户集中风险", "单一客户占收入过高导致波动", "Insilico Top5占94.4%", "高", "严重", "P0", "设置单一客户上限<20%, 平衡客户结构", "单一客户收入占比"],
    ["技术产品化风险", "实验室AI难以变成企业级产品", "多个AI创业公司踩坑", "中", "严重", "P0", "尽早投入产品/工程, 敏捷开发持续迭代", "产品化进度, 客户NPS"],
    ["销售周期过长", "材料企业决策慢, 销售周期超预期", "Citrine销售周期6-12个月", "高", "中等", "P1", "从研发部门切入, 小合同起步, ROI证明", "平均销售周期, 转化率"],
    ["竞争加剧风险", "巨头和创业公司纷纷进入赛道", "材料AI公司数量快速增长", "高", "中等", "P1", "建立数据护城河, 垂直深耕, 生态绑定", "市场份额, 客户流失率"],
    ["现金流风险", "烧钱过快, 融资不及时", "多个AI公司倒闭案例", "中", "致命", "P0", "控制burn rate, 项目制补充收入, ≥18个月现金", "现金流月数, burn rate"],
    ["人才流失风险", "AI+材料跨学科人才稀缺", "行业普遍问题", "中", "严重", "P1", "有竞争力薪酬, 文化建设, 人才梯队", "核心员工流失率"],
    ["数据可获得性", "材料数据质量差, 难以获取", "行业共性问题", "高", "中等", "P1", "与客户共创数据, 标准化工具, 自建数据集", "数据集规模, 数据质量"],
]

df7 = pd.DataFrame(risk_data[1:], columns=risk_data[0])
for r in dataframe_to_rows(df7, index=False, header=True):
    ws7.append(r)
style_worksheet(ws7, df7)

# 保存文件
output_path = "/Users/mettlyz/.openclaw/workspace/output/task-2017/AI4S公司对比数据_20260426.xlsx"
wb.save(output_path)
# print(f"Excel文件已保存至: {output_path}")
# print(f"共创建 {len(wb.sheetnames)} 个工作表:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    # print(f"  {i}. {sheet_name}")
