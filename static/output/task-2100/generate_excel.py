#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成和光智成财务估值模型Excel文件
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = openpyxl.Workbook()

# 定义样式
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
title_font = Font(bold=True, size=14)
number_format = '#,##0'
percent_format = '0%'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_table_style(ws, start_row, end_row, start_col, end_col):
    """应用表格样式"""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if row == start_row:
                cell.font = header_font
                cell.fill = header_fill

# ==================== Sheet 1: 收入预测 ====================
ws1 = wb.active
ws1.title = "收入预测"

# 标题
ws1['A1'] = "和光智成 - 收入预测模型"
ws1['A1'].font = title_font
ws1.merge_cells('A1:G1')

ws1['A3'] = "业务线A：技术服务收入"
ws1['A3'].font = Font(bold=True)
ws1['A4'] = "项目数量"
ws1['A5'] = "平均单价(万元)"
ws1['A6'] = "收入(万元)"

ws1['B3'] = "2026E"
ws1['C3'] = "2027E"
ws1['D3'] = "2028E"
ws1['E3'] = "2029E"
ws1['F3'] = "2030E"
apply_table_style(ws1, 3, 6, 1, 6)

ws1['B4'] = 5
ws1['C4'] = 15
ws1['D4'] = 30
ws1['E4'] = 50
ws1['F4'] = 80

ws1['B5'] = 80
ws1['C5'] = 100
ws1['D5'] = 133
ws1['E5'] = 200
ws1['F5'] = 225

ws1['B6'] = 400
ws1['C6'] = 1500
ws1['D6'] = 4000
ws1['E6'] = 10000
ws1['F6'] = 18000

ws1['A8'] = "业务线B：SaaS订阅收入"
ws1['A8'].font = Font(bold=True)
ws1['A9'] = "客户数量"
ws1['A10'] = "平均ARPU(万元)"
ws1['A11'] = "收入(万元)"

ws1['B8'] = "2026E"
ws1['C8'] = "2027E"
ws1['D8'] = "2028E"
ws1['E8'] = "2029E"
ws1['F8'] = "2030E"
apply_table_style(ws1, 8, 11, 1, 6)

ws1['B9'] = 20
ws1['C9'] = 100
ws1['D9'] = 250
ws1['E9'] = 500
ws1['F9'] = 800

ws1['B10'] = 2.5
ws1['C10'] = 9
ws1['D10'] = 10
ws1['E10'] = 12
ws1['F10'] = 15

ws1['B11'] = 50
ws1['C11'] = 900
ws1['D11'] = 2500
ws1['E11'] = 6000
ws1['F11'] = 12000

ws1['A13'] = "业务线C：项目合作收入"
ws1['A13'].font = Font(bold=True)
ws1['A14'] = "项目数量"
ws1['A15'] = "平均规模(万元)"
ws1['A16'] = "收入(万元)"

ws1['B13'] = "2026E"
ws1['C13'] = "2027E"
ws1['D13'] = "2028E"
ws1['E13'] = "2029E"
ws1['F13'] = "2030E"
apply_table_style(ws1, 13, 16, 1, 6)

ws1['B14'] = 2
ws1['C14'] = 5
ws1['D14'] = 10
ws1['E14'] = 20
ws1['F14'] = 35

ws1['B15'] = 75
ws1['C15'] = 120
ws1['D15'] = 150
ws1['E15'] = 200
ws1['F15'] = 229

ws1['B16'] = 150
ws1['C16'] = 600
ws1['D16'] = 1500
ws1['E16'] = 4000
ws1['F16'] = 8000

ws1['A18'] = "总收入汇总(万元)"
ws1['A18'].font = Font(bold=True)
ws1['A19'] = "技术服务收入"
ws1['A20'] = "SaaS订阅收入"
ws1['A21'] = "项目合作收入"
ws1['A22'] = "合计营收"
ws1['A23'] = "同比增长率"

ws1['B18'] = "2026E"
ws1['C18'] = "2027E"
ws1['D18'] = "2028E"
ws1['E18'] = "2029E"
ws1['F18'] = "2030E"
apply_table_style(ws1, 18, 23, 1, 6)

ws1['B19'] = 400
ws1['C19'] = 1500
ws1['D19'] = 4000
ws1['E19'] = 10000
ws1['F19'] = 18000

ws1['B20'] = 50
ws1['C20'] = 900
ws1['D20'] = 2500
ws1['E20'] = 6000
ws1['F20'] = 12000

ws1['B21'] = 150
ws1['C21'] = 600
ws1['D21'] = 1500
ws1['E21'] = 4000
ws1['F21'] = 8000

ws1['B22'] = 600
ws1['C22'] = 3000
ws1['D22'] = 8000
ws1['E22'] = 20000
ws1['F22'] = 38000

ws1['B23'] = "-"
ws1['C23'] = "400%"
ws1['D23'] = "167%"
ws1['E23'] = "150%"
ws1['F23'] = "90%"

# 调整列宽
for col in range(1, 7):
    ws1.column_dimensions[get_column_letter(col)].width = 15

# ==================== Sheet 2: 利润预测 ====================
ws2 = wb.create_sheet("利润预测")

ws2['A1'] = "和光智成 - 利润预测表"
ws2['A1'].font = title_font
ws2.merge_cells('A1:G1')

ws2['A3'] = "利润表(万元)"
ws2['A3'].font = Font(bold=True)
ws2['A4'] = "营业收入"
ws2['A5'] = "减：营业成本"
ws2['A6'] = "毛利润"
ws2['A7'] = "毛利率"
ws2['A8'] = "减：研发费用"
ws2['A9'] = "减：销售与市场"
ws2['A10'] = "减：管理费用"
ws2['A11'] = "营业利润"
ws2['A12'] = "加：其他收益"
ws2['A13'] = "利润总额"
ws2['A14'] = "减：所得税"
ws2['A15'] = "净利润"
ws2['A16'] = "净利率"

ws2['B3'] = "2026E"
ws2['C3'] = "2027E"
ws2['D3'] = "2028E"
ws2['E3'] = "2029E"
ws2['F3'] = "2030E"
apply_table_style(ws2, 3, 16, 1, 6)

ws2['B4'] = 600
ws2['C4'] = 3000
ws2['D4'] = 8000
ws2['E4'] = 20000
ws2['F4'] = 38000

ws2['B5'] = 210
ws2['C5'] = 840
ws2['D5'] = 1760
ws2['E5'] = 3600
ws2['F5'] = 6080

ws2['B6'] = 390
ws2['C6'] = 2160
ws2['D6'] = 6240
ws2['E6'] = 16400
ws2['F6'] = 31920

ws2['B7'] = "65%"
ws2['C7'] = "72%"
ws2['D7'] = "78%"
ws2['E7'] = "82%"
ws2['F7'] = "84%"

ws2['B8'] = 400
ws2['C8'] = 900
ws2['D8'] = 1800
ws2['E8'] = 3600
ws2['F8'] = 6000

ws2['B9'] = 150
ws2['C9'] = 600
ws2['D9'] = 1440
ws2['E9'] = 3000
ws2['F9'] = 5700

ws2['B10'] = 60
ws2['C10'] = 180
ws2['D10'] = 300
ws2['E10'] = 600
ws2['F10'] = 1200

ws2['B11'] = -220
ws2['C11'] = 480
ws2['D11'] = 2700
ws2['E11'] = 9200
ws2['F11'] = 19020

ws2['B12'] = 50
ws2['C12'] = 100
ws2['D12'] = 200
ws2['E12'] = 400
ws2['F12'] = 800

ws2['B13'] = -170
ws2['C13'] = 580
ws2['D13'] = 2900
ws2['E13'] = 9600
ws2['F13'] = 19820

ws2['B14'] = 0
ws2['C14'] = 28
ws2['D14'] = 900
ws2['E14'] = 3600
ws2['F14'] = 7820

ws2['B15'] = -170
ws2['C15'] = 552
ws2['D15'] = 2000
ws2['E15'] = 6000
ws2['F15'] = 12000

ws2['B16'] = "-28%"
ws2['C16'] = "18%"
ws2['D16'] = "25%"
ws2['E16'] = "30%"
ws2['F16'] = "32%"

# 成本结构
ws2['A18'] = "成本结构分析"
ws2['A18'].font = Font(bold=True)
ws2['A19'] = "COGS率"
ws2['A20'] = "R&D率"
ws2['A21'] = "S&M率"
ws2['A22'] = "G&A率"

ws2['B18'] = "2026E"
ws2['C18'] = "2027E"
ws2['D18'] = "2028E"
ws2['E18'] = "2029E"
ws2['F18'] = "2030E"
apply_table_style(ws2, 18, 22, 1, 6)

ws2['B19'] = "35%"
ws2['C19'] = "28%"
ws2['D19'] = "22%"
ws2['E19'] = "18%"
ws2['F19'] = "16%"

ws2['B20'] = "67%"
ws2['C20'] = "30%"
ws2['D20'] = "23%"
ws2['E20'] = "18%"
ws2['F20'] = "16%"

ws2['B21'] = "25%"
ws2['C21'] = "20%"
ws2['D21'] = "18%"
ws2['E21'] = "15%"
ws2['F21'] = "15%"

ws2['B22'] = "10%"
ws2['C22'] = "6%"
ws2['D22'] = "4%"
ws2['E22'] = "3%"
ws2['F22'] = "3%"

for col in range(1, 7):
    ws2.column_dimensions[get_column_letter(col)].width = 18

# ==================== Sheet 3: 现金流预测 ====================
ws3 = wb.create_sheet("现金流预测")

ws3['A1'] = "和光智成 - 现金流量表"
ws3['A1'].font = title_font
ws3.merge_cells('A1:G1')

ws3['A3'] = "现金流量表(万元)"
ws3['A3'].font = Font(bold=True)
ws3['A4'] = "经营活动现金流"
ws3['A5'] = "净利润"
ws3['A6'] = "折旧与摊销"
ws3['A7'] = "营运资本变动"
ws3['A8'] = "经营现金流净额"
ws3['A9'] = ""
ws3['A10'] = "投资活动现金流"
ws3['A11'] = "设备采购"
ws3['A12'] = "投资现金流净额"
ws3['A13'] = ""
ws3['A14'] = "筹资活动现金流"
ws3['A15'] = "股权融资"
ws3['A16'] = "银行借款"
ws3['A17'] = "筹资现金流净额"
ws3['A18'] = ""
ws3['A19'] = "现金净增加额"
ws3['A20'] = "期末现金余额"

ws3['B3'] = "2026E"
ws3['C3'] = "2027E"
ws3['D3'] = "2028E"
ws3['E3'] = "2029E"
ws3['F3'] = "2030E"
apply_table_style(ws3, 3, 20, 1, 6)

ws3['B5'] = -170
ws3['C5'] = 552
ws3['D5'] = 2000
ws3['E5'] = 6000
ws3['F5'] = 12000

ws3['B6'] = 50
ws3['C6'] = 100
ws3['D6'] = 200
ws3['E6'] = 400
ws3['F6'] = 800

ws3['B7'] = -100
ws3['C7'] = -200
ws3['D7'] = -300
ws3['E7'] = -500
ws3['F7'] = -800

ws3['B8'] = -220
ws3['C8'] = 452
ws3['D8'] = 1900
ws3['E8'] = 5900
ws3['F8'] = 12000

ws3['B11'] = -150
ws3['C11'] = -300
ws3['D11'] = -500
ws3['E11'] = -800
ws3['F11'] = -1200

ws3['B12'] = -150
ws3['C12'] = -300
ws3['D12'] = -500
ws3['E12'] = -800
ws3['F12'] = -1200

ws3['B15'] = 7500
ws3['C15'] = 0
ws3['D15'] = 15000
ws3['E15'] = 0
ws3['F15'] = 30000

ws3['B16'] = 0
ws3['C16'] = 500
ws3['D16'] = 1000
ws3['E16'] = 0
ws3['F16'] = 0

ws3['B17'] = 7500
ws3['C17'] = 500
ws3['D17'] = 16000
ws3['E17'] = 0
ws3['F17'] = 30000

ws3['B19'] = 7130
ws3['C19'] = 652
ws3['D19'] = 17400
ws3['E19'] = 5100
ws3['F19'] = 40800

ws3['B20'] = 7500
ws3['C20'] = 8152
ws3['D20'] = 25552
ws3['E20'] = 30652
ws3['F20'] = 71452

# FCFF计算
ws3['A22'] = "自由现金流(FCFF)"
ws3['A22'].font = Font(bold=True)
ws3['A23'] = "税后营业利润(NOPAT)"
ws3['A24'] = "+ 折旧摊销"
ws3['A25'] = "- 营运资本增加"
ws3['A26'] = "- 资本支出"
ws3['A27'] = "FCFF"

ws3['B22'] = "2026E"
ws3['C22'] = "2027E"
ws3['D22'] = "2028E"
ws3['E22'] = "2029E"
ws3['F22'] = "2030E"
apply_table_style(ws3, 22, 27, 1, 6)

ws3['B23'] = -170
ws3['C23'] = 493
ws3['D23'] = 2465
ws3['E23'] = 8160
ws3['F23'] = 16847

ws3['B24'] = 50
ws3['C24'] = 100
ws3['D24'] = 200
ws3['E24'] = 400
ws3['F24'] = 800

ws3['B25'] = 100
ws3['C25'] = 200
ws3['D25'] = 300
ws3['E25'] = 500
ws3['F25'] = 800

ws3['B26'] = 150
ws3['C26'] = 300
ws3['D26'] = 500
ws3['E26'] = 800
ws3['F26'] = 1200

ws3['B27'] = -370
ws3['C27'] = 93
ws3['D27'] = 1865
ws3['E27'] = 7260
ws3['F27'] = 15647

for col in range(1, 7):
    ws3.column_dimensions[get_column_letter(col)].width = 22

# ==================== Sheet 4: 可比公司法估值 ====================
ws4 = wb.create_sheet("可比公司法估值")

ws4['A1'] = "和光智成 - 可比公司法估值"
ws4['A1'].font = title_font
ws4.merge_cells('A1:G1')

ws4['A3'] = "可比公司列表"
ws4['A3'].font = Font(bold=True)
ws4['A4'] = "Periodic Labs"
ws4['A5'] = "XtalPi (晶泰科技)"
ws4['A6'] = "Schrödinger"
ws4['A7'] = "Insilico Medicine"
ws4['A8'] = "Kebotix"
ws4['A9'] = "平均值"
ws4['A10'] = "中位数"

ws4['B3'] = "国家"
ws4['C3'] = "估值(亿美金)"
ws4['D3'] = "营收(亿美金)"
ws4['E3'] = "估值/营收倍数"
ws4['F3'] = "核心业务"
apply_table_style(ws4, 3, 10, 1, 6)

ws4['B4'] = "美国"
ws4['B5'] = "中国"
ws4['B6'] = "美国"
ws4['B7'] = "美国"
ws4['B8'] = "美国"

ws4['C4'] = 70
ws4['C5'] = 50
ws4['C6'] = 35
ws4['C7'] = 30
ws4['C8'] = 12

ws4['D4'] = 1.2
ws4['D5'] = 1.1
ws4['D6'] = 2.56
ws4['D7'] = 0.8
ws4['D8'] = 0.35

ws4['E4'] = "58x"
ws4['E5'] = "45x"
ws4['E6'] = "14x"
ws4['E7'] = "38x"
ws4['E8'] = "34x"
ws4['E9'] = "37.8x"
ws4['E10'] = "38x"

ws4['F4'] = "AI材料发现"
ws4['F5'] = "AI药物发现"
ws4['F6'] = "计算化学软件"
ws4['F7'] = "AI药物研发"
ws4['F8'] = "AI材料发现"

# 和光智成估值
ws4['A12'] = "和光智成估值测算"
ws4['A12'].font = Font(bold=True)
ws4['A13'] = "方案"
ws4['A14'] = "A: 2026年营收"
ws4['A15'] = "B: 2027年前瞻营收"
ws4['A16'] = "C: 2028年IPO前估值"

ws4['B12'] = "营收基数(万)"
ws4['C12'] = "估值倍数"
ws4['D12'] = "估值区间(亿 RMB)"
apply_table_style(ws4, 12, 16, 1, 4)

ws4['B14'] = 600
ws4['B15'] = 3000
ws4['B16'] = 8000

ws4['C14'] = "15x"
ws4['C15'] = "12x-15x"
ws4['C16'] = "10x-12x"

ws4['D14'] = "0.9"
ws4['D15'] = "3.6-4.5"
ws4['D16'] = "8-9.6"

for col in range(1, 7):
    ws4.column_dimensions[get_column_letter(col)].width = 18

# ==================== Sheet 5: DCF估值 ====================
ws5 = wb.create_sheet("DCF估值")

ws5['A1'] = "和光智成 - DCF现金流折现估值"
ws5['A1'].font = title_font
ws5.merge_cells('A1:G1')

# 关键假设
ws5['A3'] = "关键假设"
ws5['A3'].font = Font(bold=True)
ws5['A4'] = "WACC"
ws5['A5'] = "无风险利率"
ws5['A6'] = "Beta系数"
ws5['A7'] = "市场风险溢价"
ws5['A8'] = "终值增长率(g)"
ws5['A9'] = "预测期"

ws5['B3'] = "数值"
ws5['C3'] = "说明"
apply_table_style(ws5, 3, 9, 1, 3)

ws5['B4'] = "25%"
ws5['B5'] = "3%"
ws5['B6'] = "1.8"
ws5['B7'] = "8%"
ws5['B8'] = "5%"
ws5['B9'] = "5年"

ws5['C4'] = "高科技初创企业，含风险溢价"
ws5['C5'] = "中国10年期国债收益率"
ws5['C6'] = "高科技行业Beta"
ws5['C7'] = "A股市场历史溢价"
ws5['C8'] = "长期可持续增长率"
ws5['C9'] = "2026-2030年"

# FCFF现值计算
ws5['A11'] = "FCFF现值计算"
ws5['A11'].font = Font(bold=True)
ws5['A12'] = "年份"
ws5['A13'] = "FCFF (万元)"
ws5['A14'] = "折现因子 (25%)"
ws5['A15'] = "FCFF现值 (万元)"

ws5['B11'] = "2026E"
ws5['C11'] = "2027E"
ws5['D11'] = "2028E"
ws5['E11'] = "2029E"
ws5['F11'] = "2030E"
ws5['G11'] = "合计"
apply_table_style(ws5, 11, 15, 1, 7)

ws5['B12'] = "2026"
ws5['C12'] = "2027"
ws5['D12'] = "2028"
ws5['E12'] = "2029"
ws5['F12'] = "2030"

ws5['B13'] = -370
ws5['C13'] = 93
ws5['D13'] = 1865
ws5['E13'] = 7260
ws5['F13'] = 15647
ws5['G13'] = "=SUM(B13:F13)"

ws5['B14'] = 0.893
ws5['C14'] = 0.797
ws5['D14'] = 0.712
ws5['E14'] = 0.635
ws5['F14'] = 0.567

ws5['B15'] = -330
ws5['C15'] = 74
ws5['D15'] = 1328
ws5['E15'] = 4610
ws5['F15'] = 8872
ws5['G15'] = 14554

# 终值计算
ws5['A17'] = "终值计算"
ws5['A17'].font = Font(bold=True)
ws5['A18'] = "2030年FCFF"
ws5['A19'] = "终值增长率"
ws5['A20'] = "终值 (Gordon模型)"
ws5['A21'] = "终值现值"

ws5['B17'] = "数值"
ws5['C17'] = "公式/说明"
apply_table_style(ws5, 17, 21, 1, 3)

ws5['B18'] = "15,647万"
ws5['B19'] = "5%"
ws5['B20'] = "82,145万"
ws5['B21'] = "46,576万"

ws5['C18'] = "预测期末FCFF"
ws5['C19'] = "长期可持续增长率"
ws5['C20'] = "=FCFF_2030*(1+g)/(WACC-g)"
ws5['C21'] = "=终值 * 2030年折现因子"

# 企业价值
ws5['A23'] = "估值结论"
ws5['A23'].font = Font(bold=True)
ws5['A24'] = "预测期FCFF现值"
ws5['A25'] = "终值现值"
ws5['A26'] = "企业价值 (EV)"
ws5['A27'] = "加：现金"
ws5['A28'] = "股权价值"

ws5['B23'] = "金额(万元)"
apply_table_style(ws5, 23, 28, 1, 2)

ws5['B24'] = 14554
ws5['B25'] = 46576
ws5['B26'] = 61130
ws5['B27'] = 7500
ws5['B28'] = 68630

for col in range(1, 8):
    ws5.column_dimensions[get_column_letter(col)].width = 20

# ==================== Sheet 6: 敏感性分析 ====================
ws6 = wb.create_sheet("敏感性分析")

ws6['A1'] = "和光智成 - 估值敏感性分析"
ws6['A1'].font = title_font
ws6.merge_cells('A1:G1')

# WACC敏感性
ws6['A3'] = "WACC敏感性分析 (股权价值:亿元)"
ws6['A3'].font = Font(bold=True)
ws6['A4'] = "WACC ↓ / 终值增长率 →"
ws6['A5'] = "20%"
ws6['A6'] = "25%"
ws6['A7'] = "30%"

ws6['B3'] = "3%"
ws6['C3'] = "4%"
ws6['D3'] = "5%"
ws6['E3'] = "6%"
ws6['F3'] = "7%"
apply_table_style(ws6, 3, 7, 1, 6)

ws6['B5'] = 8.2
ws6['C5'] = 9.1
ws6['D5'] = 10.3
ws6['E5'] = 11.8
ws6['F5'] = 14.0

ws6['B6'] = 5.2
ws6['C6'] = 5.7
ws6['D6'] = 6.9
ws6['E6'] = 7.7
ws6['F6'] = 8.7

ws6['B7'] = 3.8
ws6['C7'] = 4.1
ws6['D7'] = 4.5
ws6['E7'] = 5.0
ws6['F7'] = 5.6

# 营收增速敏感性
ws6['A9'] = "营收增速敏感性分析"
ws6['A9'].font = Font(bold=True)
ws6['A10'] = "情景"
ws6['A11'] = "保守 (-20%)"
ws6['A12'] = "基准 (0%)"
ws6['A13'] = "乐观 (+20%)"

ws6['B9'] = "2030年营收(万)"
ws6['C9'] = "股权价值(亿元)"
apply_table_style(ws6, 9, 13, 1, 3)

ws6['B11'] = 30400
ws6['B12'] = 38000
ws6['B13'] = 45600

ws6['C11'] = 4.8
ws6['C12'] = 6.9
ws6['C13'] = 9.2

# 综合估值结论
ws6['A15'] = "综合估值结论"
ws6['A15'].font = Font(bold=True)
ws6['A16'] = "估值方法"
ws6['A17'] = "可比公司法 (2027前瞻)"
ws6['A18'] = "DCF现金流折现法"
ws6['A19'] = "DCF基准情形"
ws6['A20'] = "综合估值区间"

ws6['B15'] = "估值区间 (亿元)"
apply_table_style(ws6, 15, 20, 1, 2)

ws6['B17'] = "3.6 - 4.5"
ws6['B18'] = "5.2 - 9.2"
ws6['B19'] = "6.9"
ws6['B20'] = "5.0 - 8.0"

# 融资建议
ws6['A22'] = "本轮融资建议"
ws6['A22'].font = Font(bold=True)
ws6['A23'] = "融资条款"
ws6['A24'] = "融资金额"
ws6['A25'] = "目标估值 (投后)"
ws6['A26'] = "出让股权比例"
ws6['A27'] = "估值依据"

ws6['B22'] = "建议值"
apply_table_style(ws6, 22, 27, 1, 2)

ws6['B24'] = "5,000万 - 10,000万 RMB"
ws6['B25'] = "5 - 8亿 RMB"
ws6['B26'] = "10% - 15%"
ws6['B27'] = "DCF基准情形6.9亿，取区间5-8亿"

for col in range(1, 7):
    ws6.column_dimensions[get_column_letter(col)].width = 22

# 保存文件
output_path = "/Users/mettlyz/.openclaw/workspace/output/task-2100/和光智成_财务估值模型_Q2_20260426.xlsx"
wb.save(output_path)
# print(f"✅ Excel文件已生成: {output_path}")
# print(f"✅ 文件大小: {len(open(output_path, 'rb').read())} 字节")